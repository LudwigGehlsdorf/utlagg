#!/usr/bin/env python3
"""Parse GENOMRÖSTAD BUDGETREVIDERING.xlsx into prisma/budget-2026.json.

Each committee sheet (AKTU01, SEX02, …) becomes a cost center; its Konto rows
become accounts and the indented rows beneath become line items. Two value
columns are read per sheet: J = "2026 REV" (the working revision) and
L = "2026" (the originally-adopted baseline).

Where a line uses the "antal × á-pris" form (cells F × H) the REV revision keeps
that structure, mapping well-known Parametrar references to named variables.
Anything else — and the whole baseline column — is stored as its computed value,
so the totals always match the spreadsheet exactly.
"""
import json
import re
import os
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

HERE = os.path.dirname(__file__)
XLSX = os.path.join(HERE, "..", "GENOMRÖSTAD BUDGETREVIDERING.xlsx")
OUT = os.path.join(HERE, "budget-2026.json")

# Parametrar rows we expose as named variables (cell Parametrar!$I$<row>).
PARAM_BY_ROW = {
    7: "PRISBASBELOPP",
    8: "TACK_POLICY",
    9: "TROJA_POLICY",
    10: "MAT",
    33: "PASLAG_OL",
    34: "PASLAG_OVRIG_JAST",
    35: "PASLAG_VIN",
    36: "PASLAG_SPRIT",
}

COL_F, COL_H, COL_J, COL_L = 6, 8, 10, 12
SKIP_B = {"Summa", "Intäkter", "Kostnader", "Resultat", ""}

wb = openpyxl.load_workbook(XLSX, data_only=False)
wbv = openpyxl.load_workbook(XLSX, data_only=True)


def fmtnum(v):
    """Format a number as a clean string ("4002", "2", "1.44", "560705.0151")."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return str(v).strip()
    if f == int(f):
        return str(int(f))
    return repr(f)


def numeric(v):
    """A cell's value as a float, or None if it isn't a plain number."""
    if v is None or isinstance(v, bool):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def param_ref(formula):
    """Map '=Parametrar!$I$8' → 'TACK_POLICY', else None."""
    if not isinstance(formula, str):
        return None
    m = re.match(r"^=\s*Parametrar!\$?I\$?(\d+)\s*$", formula)
    if not m:
        return None
    return PARAM_BY_ROW.get(int(m.group(1)))


def cell_term(sheet_f, sheet_v, r, c):
    """Resolve a quantity/á-pris cell to an expression string.

    A Parametrar reference becomes its variable name; anything else becomes the
    computed numeric value. Returns None when the cell is empty.
    """
    raw = sheet_f.cell(r, c).value
    if raw is None or raw == "":
        return None
    if isinstance(raw, str) and raw.startswith("="):
        var = param_ref(raw)
        if var:
            return var
        return fmtnum(sheet_v.cell(r, c).value)
    if isinstance(raw, ArrayFormula):
        return fmtnum(sheet_v.cell(r, c).value)
    return fmtnum(raw)


def value_expr(sheet_f, sheet_v, r, c):
    """A standalone amount cell → variable name if it's a plain Parametrar ref,
    otherwise the computed value as a string."""
    raw = sheet_f.cell(r, c).value
    if isinstance(raw, str) and raw.startswith("="):
        var = param_ref(raw)
        if var:
            return var
    val = sheet_v.cell(r, c).value
    return fmtnum(val) if val is not None else "0"


def committee_map():
    """Build {sheetCode: committeeName} from the Rambudget overview."""
    ws = wb["Rambudget"]
    sheet_names = set(wb.sheetnames)
    out, current = {}, None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if not isinstance(a, str) or not a.strip():
            continue
        a = a.strip()
        if a in sheet_names:
            if current:
                out[a] = current
        elif a.startswith("Totalt") or a in ("Beteckning", "Allmän info", "Totalt"):
            continue
        else:
            current = a  # a committee header row
    return out


def parse_variables():
    ws = wbv["Parametrar"]
    out = []
    for row, name in PARAM_BY_ROW.items():
        out.append({
            "name": name,
            "rev": fmtnum(ws.cell(row, COL_J - 1).value) or "0",   # column I
            "orig": fmtnum(ws.cell(row, COL_L - 1).value) or "0",  # column K
        })
    return out


def parse_sheet(code):
    sf, sv = wb[code], wbv[code]
    name = sf.cell(1, 2).value or code
    # find the "Konto" header row
    start = None
    for r in range(1, sf.max_row + 1):
        if str(sf.cell(r, 1).value).strip() == "Konto":
            start = r + 1
            break
    if start is None:
        start = 11

    accounts, by_code, current = [], {}, None
    for r in range(start, sf.max_row + 1):
        a = sf.cell(r, 1).value
        b = sf.cell(r, 2).value
        b_str = str(b).strip() if b is not None else ""

        if a is not None and str(a).strip() != "" and b_str not in ("Intäkter", "Kostnader", "Resultat"):
            # account header
            acode = fmtnum(a) if isinstance(a, (int, float)) else str(a).strip()
            if acode in by_code:
                current = by_code[acode]
            else:
                current = {"code": acode, "name": b_str, "lineItems": []}
                by_code[acode] = current
                accounts.append(current)
            continue

        if b_str in SKIP_B or current is None:
            continue

        # line item. The amount is always column J (REV) / L (baseline) — these
        # are the sheet's source of truth (a row may override the formula with a
        # literal). The "antal × á-pris" form is kept for display only when the
        # row is a clean 2-factor F × H whose product matches J (no extra D
        # factor, no manual override), so the breakdown can never disagree with
        # the total.
        fv = numeric(sv.cell(r, COL_F).value)
        hv = numeric(sv.cell(r, COL_H).value)
        jv = numeric(sv.cell(r, COL_J).value)
        d_empty = sf.cell(r, 4).value in (None, "")
        structured = (fv is not None and hv is not None and d_empty
                      and jv is not None and abs(fv * hv - jv) < 1.0)
        if structured:
            rev = {"quantity": cell_term(sf, sv, r, COL_F),
                   "unitPrice": cell_term(sf, sv, r, COL_H),
                   "expression": value_expr(sf, sv, r, COL_J)}
        else:
            rev = {"quantity": None, "unitPrice": None, "expression": value_expr(sf, sv, r, COL_J)}
        orig = {"quantity": None, "unitPrice": None, "expression": value_expr(sf, sv, r, COL_L)}
        current["lineItems"].append({"description": b_str, "rev": rev, "orig": orig})

    # drop empty accounts (header with no lines and no name carry no signal)
    accounts = [a for a in accounts if a["lineItems"]]
    return {"code": code, "name": str(name).strip(), "lineItems_unused": False, "accounts": accounts}


def main():
    cm = committee_map()
    skip = {"Rambudget", "Parametrar"}
    cost_centers = []
    for code in wb.sheetnames:
        if code in skip:
            continue
        cc = parse_sheet(code)
        cc["committee"] = cm.get(code)
        del cc["lineItems_unused"]
        cost_centers.append(cc)

    data = {"year": 2026, "variables": parse_variables(), "costCenters": cost_centers}
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=1)

    n_acc = sum(len(c["accounts"]) for c in cost_centers)
    n_li = sum(len(a["lineItems"]) for c in cost_centers for a in c["accounts"])
    print(f"Wrote {OUT}: {len(cost_centers)} cost centers, {n_acc} accounts, {n_li} line items, {len(data['variables'])} variables")


if __name__ == "__main__":
    main()
